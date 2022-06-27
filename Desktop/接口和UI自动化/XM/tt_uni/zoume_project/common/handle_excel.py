"""
封装的需求
1、封装一个可以读取任意的excel文件的方法，可以指定读取的表单
2、数据写入：


"""

import openpyxl


class HandleExcel:

    def __init__(self, filename, sheetname):
        """

        :param filename:excel文件名名（路径）
        :param sheetname:表单名
        """
        self.filename = filename
        self.sheetname = sheetname


    def read_data(self):
        """ 读取excel的数据"""
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]

        # rows:按行获取表单中所有的格子,每一行的格子放到一个元祖中
        res = list(sh.rows)

        # 推导式获取excel中第一行的数据
        title = [i.value for i in res[0]]
        # print(title)

        cases = []
        # 遍历第一行以外所有的行
        for item in res[1:]:
            # 获取该行的数据
            data = [i.value for i in item]
            # 第一行的数据和当前这行数据打包为字典
            dic = dict(zip(title, data))
            # 把字典添加到cases这个列表中
            cases.append(dic)
        workbook.close()
        return cases


    def write_data(self,  row, column, value):
        """
        数据写入的方法
        :param row:  写入的行
        :param column: 写入的列
        :param value: 写入的值
        :return:
        """
        #加载工作簿对象
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        # 写入数据
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filename)

